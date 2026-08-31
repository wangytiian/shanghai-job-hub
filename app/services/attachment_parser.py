from dataclasses import dataclass
from datetime import datetime
from hashlib import sha256
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Job, ReviewLog


ROLE_HEADER_HINTS = ("岗位名称", "招聘岗位", "岗位")
MAX_ROWS_PER_SHEET = 200


@dataclass(frozen=True)
class AttachmentRoleCandidate:
    title: str
    row_number: int
    evidence: str


def _cell_text(value) -> str:
    return "" if value is None else str(value).strip()


def _find_role_column(rows: list[tuple]) -> tuple[int, int] | None:
    for row_index, row in enumerate(rows[:20]):
        values = [_cell_text(value) for value in row]
        for column_index, value in enumerate(values):
            if value in ROLE_HEADER_HINTS:
                return row_index, column_index
    return None


def parse_xlsx_role_candidates(content: bytes) -> list[AttachmentRoleCandidate]:
    """Extract only rows under an explicit role-name column; never infer roles."""
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    candidates: list[AttachmentRoleCandidate] = []
    seen_titles: set[str] = set()
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True, max_row=MAX_ROWS_PER_SHEET))
        header = _find_role_column(rows)
        if header is None:
            continue
        header_row, role_column = header
        for row_index, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
            title = _cell_text(row[role_column] if role_column < len(row) else None)
            if not title or title in ROLE_HEADER_HINTS or title in seen_titles:
                continue
            evidence = " | ".join(
                f"{_cell_text(rows[header_row][column_index])}：{_cell_text(value)}"
                for column_index, value in enumerate(row)
                if _cell_text(value) and column_index < len(rows[header_row])
            )
            if not evidence:
                continue
            candidates.append(AttachmentRoleCandidate(title=title, row_number=row_index, evidence=evidence))
            seen_titles.add(title)
    return candidates


def create_pending_child_jobs(
    session: Session,
    parent: Job,
    attachment_name: str,
    attachment_url: str,
    candidates: list[AttachmentRoleCandidate],
    operator_name: str,
) -> list[Job]:
    """Create independently reviewable roles from verified attachment rows only."""
    if parent.attachment_status != "checked":
        raise ValueError("附件尚未核验，不能拆分岗位")
    children: list[Job] = []
    for candidate in candidates:
        unique = sha256(f"{parent.id}|{attachment_url}|{candidate.row_number}|{candidate.title}".encode()).hexdigest()
        fingerprint = f"附件拆岗|{parent.id}|{unique}"
        child = session.scalar(select(Job).where(Job.fingerprint == fingerprint))
        if child is None:
            child = Job(
                fingerprint=fingerprint,
                employer_name=parent.employer_name,
                job_title=candidate.title,
                job_family="待人工分类",
                recruitment_type=parent.recruitment_type,
                location_category=parent.location_category,
                location_detail=parent.location_detail,
                target_audience="需按具体岗位判断",
                direction_tags="待人工分类",
                deadline=parent.deadline,
                official_url=parent.official_url,
                source_url=parent.source_url,
                evidence_text=f"来源附件：{attachment_name}\n附件链接：{attachment_url}\n第 {candidate.row_number} 行：{candidate.evidence}",
                quality_score=0,
                risk_flags="附件解析候选：须人工核验岗位条件后方可发布",
                is_demo=parent.is_demo,
                collected_at=datetime.now(),
                lifecycle_status="正常",
                last_change_summary="已从已核验附件拆分为待核验岗位",
                status="待核验",
                notice_type="新招聘",
                notice_type_suggestion="新招聘",
                posting_scope="single_role",
                attachment_status="checked",
                application_method=parent.application_method,
                application_contact=parent.application_contact,
                attachment_links="[]",
                parent_job_id=parent.id,
            )
            session.add(child)
            session.flush()
            session.add(
                ReviewLog(
                    job_id=child.id,
                    action="附件拆分",
                    note=f"由父公告 #{parent.id} 的 {attachment_name} 第 {candidate.row_number} 行生成",
                    operator_name=operator_name,
                )
            )
        children.append(child)
    session.commit()
    return children
